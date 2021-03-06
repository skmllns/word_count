import openpyxl
import re
import collections
import csv
import nltk

read_file = "word_count_input.xlsx"
write_file = "word_count_output.csv"
pos_to_pull = ['JJ', 'JJR', 'JJS', 'NN', 'NNS', 'NNP', 'NNPS', 'VB', 'VBD', 'VBG', 'VBN', 'VBP',
               'VBZ']
               
words_to_avoid = ['be', 'had', 'is', 'am', 'was', 'has', 'are', 'do']

headers = ["id", "class", "enrollment", "gender", "word"]
wb = openpyxl.load_workbook(read_file)
sheet = wb.get_sheet_by_name('Sheet1')
row_count = sheet.get_highest_row() - 1
     
with open(write_file, 'wb') as f: 
   writer = csv.writer(f)
   writer.writerow(headers)
   for row_num in range(1, row_count):
      year = sheet.cell(row=row_num, column=0).value
      enrollment = sheet.cell(row=row_num, column=1).value
      gender = sheet.cell(row=row_num, column=2).value
      words = re.findall(r'\w+\'*\w*', sheet.cell(row=row_num, column=3).value.lower())
      tagged_words = nltk.pos_tag(words)
      for word in tagged_words:
         if word[1] in pos_to_pull and word[0] not in words_to_avoid:
            writer.writerow([row_num, year, enrollment, gender, word[0]])

      
